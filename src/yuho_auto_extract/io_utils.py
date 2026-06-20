from __future__ import annotations

import csv
import html
import importlib.util
import json
import math
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import yaml


Record = Dict[str, Any]
BLANKISH_STRINGS = {"", "nan", "none", "null", "<na>"}


def project_root_from(path: Optional[str] = None) -> Path:
    if path:
        return Path(path).resolve()
    return Path.cwd().resolve()


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def read_yaml(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def write_yaml(path: Path, data: Dict[str, Any]) -> None:
    ensure_parent(path)
    with path.open("w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False)


def read_jsonl(path: Path) -> List[Record]:
    if not path.exists():
        return []
    rows: List[Record] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def write_jsonl(path: Path, rows: Iterable[Record]) -> None:
    ensure_parent(path)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def read_csv_records(path: Path) -> List[Record]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [dict(row) for row in csv.DictReader(f)]


def write_csv_records(path: Path, rows: Iterable[Record]) -> None:
    rows_list = list(rows)
    ensure_parent(path)
    if not rows_list:
        path.write_text("", encoding="utf-8")
        return
    fieldnames: List[str] = []
    for row in rows_list:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows_list:
            writer.writerow({key: _stringify_cell(row.get(key)) for key in fieldnames})


def _stringify_cell(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if is_blankish(value):
        return ""
    return value


def is_blankish(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    try:
        return str(value).strip().lower() in BLANKISH_STRINGS
    except Exception:
        return False


def read_table(path: Path) -> List[Record]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_records(path)
    if suffix == ".jsonl":
        return read_jsonl(path)
    if suffix == ".json":
        return json.loads(path.read_text(encoding="utf-8"))
    if suffix == ".yml" or suffix == ".yaml":
        data = read_yaml(path)
        if isinstance(data, list):
            return data
        raise ValueError(f"{path} is YAML but does not contain a list")
    if suffix == ".xlsx":
        return _read_xlsx_records(path)
    if suffix == ".parquet":
        try:
            import pandas as pd  # type: ignore
        except ImportError as exc:
            raise RuntimeError("Reading parquet requires pandas and pyarrow") from exc
        return pd.read_parquet(path).to_dict(orient="records")
    raise ValueError(f"Unsupported table format: {path}")


def write_table(path: Path, rows: Iterable[Record]) -> Path:
    rows_list = list(rows)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        write_csv_records(path, rows_list)
        return path
    if suffix == ".jsonl":
        write_jsonl(path, rows_list)
        return path
    if suffix == ".json":
        ensure_parent(path)
        path.write_text(json.dumps(rows_list, ensure_ascii=False, indent=2), encoding="utf-8")
        return path
    if suffix == ".xlsx":
        try:
            _write_xlsx_records(path, rows_list)
            return path
        except ImportError:
            fallback = path.with_suffix(".csv")
            write_csv_records(fallback, rows_list)
            return fallback
    if suffix == ".parquet":
        try:
            import pandas as pd  # type: ignore
        except ImportError:
            fallback = path.with_suffix(".csv")
            write_csv_records(fallback, rows_list)
            return fallback
        ensure_parent(path)
        _dataframe_for_parquet(pd, rows_list).to_parquet(path, index=False)
        return path
    raise ValueError(f"Unsupported table format: {path}")


def _dataframe_for_parquet(pd: Any, rows: List[Record]) -> Any:
    df = pd.DataFrame(rows)
    for column in df.columns:
        if str(df[column].dtype) != "object":
            continue
        values = [value for value in df[column].tolist() if not is_blankish(value)]
        value_types = {type(value) for value in values}
        has_structured = any(isinstance(value, (dict, list)) for value in values)
        if has_structured or len(value_types) > 1:
            df[column] = df[column].map(_stringify_cell_for_parquet)
    return df


def _stringify_cell_for_parquet(value: Any) -> str:
    value = _stringify_cell(value)
    if is_blankish(value):
        return ""
    return str(value)


def _read_xlsx_records(path: Path) -> List[Record]:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Reading xlsx requires openpyxl; use the CSV templates or install project dependencies") from exc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    records: List[Record] = []
    for row in rows[1:]:
        records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return records


def _write_xlsx_records(path: Path, rows: List[Record]) -> None:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        _write_minimal_xlsx(path, rows)
        return
    ensure_parent(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    if rows:
        headers: List[str] = []
        for row in rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)
        ws.append(headers)
        for row in rows:
            ws.append([_stringify_cell(row.get(key)) for key in headers])
    wb.save(path)


def _write_minimal_xlsx(path: Path, rows: List[Record]) -> None:
    ensure_parent(path)
    headers: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)
    sheet_rows = [headers] + [[_stringify_cell(row.get(key)) for key in headers] for row in rows]
    worksheet = _worksheet_xml(sheet_rows)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", _CONTENT_TYPES_XML)
        zf.writestr("_rels/.rels", _RELS_XML)
        zf.writestr("xl/workbook.xml", _WORKBOOK_XML)
        zf.writestr("xl/_rels/workbook.xml.rels", _WORKBOOK_RELS_XML)
        zf.writestr("xl/worksheets/sheet1.xml", worksheet)


def _worksheet_xml(rows: List[List[Any]]) -> str:
    body = []
    for row_index, row in enumerate(rows, start=1):
        cells = []
        for col_index, value in enumerate(row, start=1):
            ref = f"{_column_name(col_index)}{row_index}"
            text = html.escape("" if value is None else str(value), quote=False)
            cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{text}</t></is></c>')
        body.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{''.join(body)}</sheetData>"
        "</worksheet>"
    )


def _column_name(index: int) -> str:
    name = ""
    while index:
        index, rem = divmod(index - 1, 26)
        name = chr(65 + rem) + name
    return name


_CONTENT_TYPES_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>"""

_RELS_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""

_WORKBOOK_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
  xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets>
</workbook>"""

_WORKBOOK_RELS_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>"""


def prefer_existing_table(base_path: Path) -> Path:
    if base_path.suffix.lower() == ".xlsx":
        csv_path = base_path.with_suffix(".csv")
        if csv_path.exists() and importlib.util.find_spec("openpyxl") is None:
            return csv_path
    if base_path.suffix.lower() == ".parquet":
        csv_path = base_path.with_suffix(".csv")
        jsonl_path = base_path.with_suffix(".jsonl")
        if csv_path.exists() and (importlib.util.find_spec("pandas") is None or importlib.util.find_spec("pyarrow") is None):
            return csv_path
        if jsonl_path.exists() and (importlib.util.find_spec("pandas") is None or importlib.util.find_spec("pyarrow") is None):
            return jsonl_path
    if base_path.exists():
        return base_path
    if base_path.suffix.lower() == ".xlsx":
        csv_path = base_path.with_suffix(".csv")
        if csv_path.exists():
            return csv_path
    if base_path.suffix.lower() == ".parquet":
        csv_path = base_path.with_suffix(".csv")
        if csv_path.exists():
            return csv_path
        jsonl_path = base_path.with_suffix(".jsonl")
        if jsonl_path.exists():
            return jsonl_path
    return base_path
