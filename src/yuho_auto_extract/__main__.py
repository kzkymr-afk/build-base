from __future__ import annotations

import argparse
import json
import os
import uuid
import zipfile
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List

from .ai_bundle import build_ai_bundle
from .analysis_builder import build_analysis_dataset
from .config_loader import field_map, load_pipeline_config
from .document_resolver import resolve_target_documents
from .edinet_db import build_edinet_db, extract_from_edinet_db, split_local_review_rows
from .exporter import apply_review_decisions, build_source_audit, build_wide_values, filter_exportable_rows
from .io_utils import prefer_existing_table, read_table, write_table
from .logging_utils import configure_logging
from .normalizer import normalize_extraction
from .review_queue import build_review_queue
from .run_reporter import build_run_report, write_run_report
from .section_locator import locate_candidate_blocks
from .services.algorithm_audit import build_algorithm_audit_bundle
from .services.manual_technicians import import_manual_technicians
from .validator import attach_validation_status, validate_records
from .xbrl_csv_parser import extract_xbrl_csv_long
from .xbrl_fact_store import build_xbrl_fact_store, compare_xbrl_fact_store, extract_from_xbrl_fact_store


def main(argv: List[str] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    configure_logging()
    root = Path(args.root).resolve()
    _load_env_file(root)
    if not hasattr(args, "handler"):
        parser.print_help()
        return 2
    return args.handler(root, args)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="python -m yuho_auto_extract")
    parser.add_argument("--root", default=".", help="Project root containing config/ and data/")
    sub = parser.add_subparsers(dest="command")

    p = sub.add_parser("index")
    p.add_argument("--start-date", required=True)
    p.add_argument("--end-date", required=True)
    p.add_argument("--merge", action="store_true", help="Merge fetched rows into the existing document_index instead of replacing it.")
    p.set_defaults(handler=cmd_index)

    p = sub.add_parser("index-annual")
    p.add_argument("--fiscal-years", nargs="*", type=int)
    p.set_defaults(handler=cmd_index_annual)

    p = sub.add_parser("resolve")
    p.add_argument("--fiscal-years", nargs="*", type=int)
    p.add_argument("--period-type", choices=["annual", "semiannual_h1", "all"], default="annual")
    p.add_argument("--merge", action="store_true", help="Merge resolved rows into the existing target_documents instead of replacing it.")
    p.set_defaults(handler=cmd_resolve)

    p = sub.add_parser("download")
    p.add_argument("--target-documents", default="data/intermediate/target_documents.parquet")
    p.set_defaults(handler=cmd_download)

    p = sub.add_parser("build-edinet-db")
    p.add_argument("--db", default="data/intermediate/edinet.db")
    p.set_defaults(handler=cmd_build_edinet_db)

    p = sub.add_parser("extract-from-db")
    p.add_argument("--db", default="data/intermediate/edinet.db")
    p.add_argument("--output", default="data/intermediate/db_extracted_long.csv")
    p.add_argument("--no-pipeline", action="store_true")
    p.add_argument("--period-type", choices=["annual", "semiannual_h1", "all"], default="annual")
    p.set_defaults(handler=cmd_extract_from_db)

    p = sub.add_parser("run-local")
    p.add_argument("--db", default="data/intermediate/edinet.db")
    p.set_defaults(handler=cmd_run_local)

    p = sub.add_parser("run-all")
    p.add_argument("--db", default="data/intermediate/edinet.db")
    p.add_argument("--fiscal-years", nargs="*", type=int)
    p.set_defaults(handler=cmd_run_all)

    p = sub.add_parser("extract-xbrl")
    p.add_argument("--period-type", choices=["annual", "semiannual_h1", "all"], default="annual")
    p.set_defaults(handler=cmd_extract_xbrl)
    p = sub.add_parser("build-xbrl-fact-store")
    p.add_argument("--doc-id", default="")
    p.add_argument("--company-year-id", default="")
    p.add_argument("--merge", action="store_true", help="Replace selected facts while preserving other fact-store rows.")
    p.set_defaults(handler=cmd_build_xbrl_fact_store)
    p = sub.add_parser("extract-from-xbrl-fact-store")
    p.add_argument("--output", default="data/intermediate/xbrl_fact_store_extracted_long.csv")
    p.add_argument("--write-pipeline", action="store_true")
    p.add_argument("--company-year-id", default="")
    p.set_defaults(handler=cmd_extract_from_xbrl_fact_store)
    p = sub.add_parser("compare-xbrl-fact-store")
    p.add_argument("--old", default="data/intermediate/xbrl_extracted_long.parquet")
    p.add_argument("--new", default="data/intermediate/xbrl_fact_store_extracted_long.csv")
    p.add_argument("--output", default="data/reports/xbrl_fact_store_compare.csv")
    p.set_defaults(handler=cmd_compare_xbrl_fact_store)
    sub.add_parser("locate-sections").set_defaults(handler=cmd_locate_sections)

    sub.add_parser("normalize").set_defaults(handler=cmd_normalize)
    sub.add_parser("validate").set_defaults(handler=cmd_validate)
    sub.add_parser("corroborate").set_defaults(handler=cmd_corroborate)

    p = sub.add_parser("build-review-queue")
    p.add_argument("--existing-excel", default="")
    p.set_defaults(handler=cmd_build_review_queue)

    sub.add_parser("split-local-review").set_defaults(handler=cmd_split_local_review)
    p = sub.add_parser("import-manual-technicians")
    p.add_argument("--note-path", default="")
    p.set_defaults(handler=cmd_import_manual_technicians)

    p = sub.add_parser("export-final")
    p.add_argument("--reviewed", default="data/review/review_resolved.csv")
    p.set_defaults(handler=cmd_export_final)

    sub.add_parser("build-analysis").set_defaults(handler=cmd_build_analysis)
    sub.add_parser("report").set_defaults(handler=cmd_report)
    sub.add_parser("build-ai-bundle").set_defaults(handler=cmd_build_ai_bundle)
    sub.add_parser("build-algorithm-audit").set_defaults(handler=cmd_build_algorithm_audit)
    sub.add_parser("build-corroboration-report").set_defaults(handler=cmd_build_corroboration_report)
    sub.add_parser("audit-findings").set_defaults(handler=cmd_audit_findings)
    sub.add_parser("golden-freeze").set_defaults(handler=cmd_golden_freeze)
    sub.add_parser("backfill-semantics").set_defaults(handler=cmd_backfill_semantics)
    sub.add_parser("semantics-coverage").set_defaults(handler=cmd_semantics_coverage)
    p = sub.add_parser("ai-map-observed-items")
    p.add_argument("--tier", default="bulk", choices=["bulk", "hard", "audit"])
    p.add_argument("--limit", type=int, default=50)
    p.add_argument("--taxonomy-kind", dest="taxonomy_kind", default=None,
                   choices=["jppfs", "jpcrp", "ifrs", "extension", "local"])
    p.add_argument("--dry-run", dest="dry_run", action="store_true", default=True)
    p.add_argument("--no-dry-run", dest="dry_run", action="store_false")
    p.set_defaults(handler=cmd_ai_map_observed_items)
    p = sub.add_parser("ai-verify-mappings")
    p.set_defaults(handler=cmd_ai_verify_mappings)
    p = sub.add_parser("promote-ai-mappings")
    p.add_argument("--confirm-verified-maps", dest="confirm_verified_maps", action="store_true", default=None)
    p.add_argument("--adopt-new-concepts", dest="adopt_new_concepts", action="store_true", default=None)
    p.add_argument("--dry-run", dest="dry_run", action="store_true", default=True)
    p.add_argument("--no-dry-run", dest="dry_run", action="store_false")
    p.set_defaults(handler=cmd_promote_ai_mappings)
    p = sub.add_parser("regression-check")
    p.add_argument("--mode", choices=["light", "full"], default="light")
    p.set_defaults(handler=cmd_regression_check)
    p = sub.add_parser("enrich-field-definition")
    p.add_argument("--dry-run", dest="dry_run", action="store_true", default=True)
    p.add_argument("--apply", dest="dry_run", action="store_false")
    p.set_defaults(handler=cmd_enrich_field_definition)
    p = sub.add_parser("automation-status")
    p.add_argument("--fiscal-year", type=int)
    p.set_defaults(handler=cmd_automation_status)
    p = sub.add_parser("roll-forward-year")
    p.add_argument("--fiscal-year", required=True, type=int)
    p.add_argument("--dry-run", action="store_true")
    p.set_defaults(handler=cmd_roll_forward_year)
    p = sub.add_parser("annual-refresh")
    p.add_argument("--fiscal-year", type=int)
    p.add_argument("--force", action="store_true")
    p.add_argument("--dry-run", action="store_true")
    p.set_defaults(handler=cmd_annual_refresh)
    sub.add_parser("stock-status").set_defaults(handler=cmd_stock_status)
    p = sub.add_parser("stock-refresh")
    p.add_argument("--force", action="store_true")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--start-date", default="")
    p.add_argument("--end-date", default="")
    p.set_defaults(handler=cmd_stock_refresh)
    sub.add_parser("factbook-status").set_defaults(handler=cmd_factbook_status)
    p = sub.add_parser("factbook-refresh")
    p.add_argument("--force", action="store_true")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--use-ai", action="store_true")
    p.add_argument("--ai-tier", default="bulk")
    p.add_argument("--company", action="append", default=[], help="Limit refresh to a company id. Can be specified multiple times.")
    p.add_argument("--source", action="append", default=[], help="Limit refresh to a source id or source_dataset_id. Can be specified multiple times.")
    p.add_argument("--fiscal-year", action="append", default=[], help="Limit parsed source documents to a fiscal year. Can be specified multiple times.")
    p.set_defaults(handler=cmd_factbook_refresh)
    sub.add_parser("factbook-validate").set_defaults(handler=cmd_factbook_validate)
    sub.add_parser("factbook-coverage").set_defaults(handler=cmd_factbook_coverage)
    sub.add_parser("init-xlsx").set_defaults(handler=cmd_init_xlsx)
    p = sub.add_parser("infer-sources-from-confirmed")
    p.add_argument(
        "--fields",
        default=",".join(["building_orders_total", "completed_building", "backlog_building_next"]),
        help="Comma-separated field_ids to run source inference for.",
    )
    p.set_defaults(handler=cmd_infer_sources_from_confirmed)
    p = sub.add_parser("apply-source-inference")
    p.add_argument(
        "--fields",
        default=",".join(["building_orders_total", "completed_building", "backlog_building_next"]),
        help="Comma-separated field_ids to run source inference promotion for.",
    )
    p.add_argument("--dry-run", dest="dry_run", action="store_true", default=True)
    p.add_argument("--apply", dest="dry_run", action="store_false")
    p.set_defaults(handler=cmd_apply_source_inference)
    p = sub.add_parser("web")
    p.add_argument("--host", default="127.0.0.1")
    p.add_argument("--port", default=8765, type=int)
    p.add_argument("--reload", action="store_true")
    p.set_defaults(handler=cmd_web)
    return parser


def cmd_index(root: Path, args: argparse.Namespace) -> int:
    from .edinet_client import EdinetClient, save_index_response

    cfg = load_pipeline_config(root)
    edinet_cfg = cfg.model_config.get("edinet", {})
    client = EdinetClient(
        base_url=edinet_cfg.get("base_url", "https://api.edinet-fsa.go.jp/api/v2"),
        timeout=int(edinet_cfg.get("request_timeout_seconds", 30)),
        retry_count=int(edinet_cfg.get("retry_count", 3)),
        retry_backoff_seconds=int(edinet_cfg.get("retry_backoff_seconds", 2)),
    )
    rows = _collect_index_rows(client, root, _date_range(_parse_date(args.start_date), _parse_date(args.end_date)))
    if getattr(args, "merge", False):
        existing_path = prefer_existing_table(root / "data" / "intermediate" / "document_index.parquet")
        existing = read_table(existing_path) if existing_path.exists() else []
        rows = _merge_rows_by_key(existing, rows, "docID")
    written = write_table(root / "data" / "intermediate" / "document_index.parquet", rows)
    print(f"wrote {written} rows={len(rows)}")
    return 0


def cmd_index_annual(root: Path, args: argparse.Namespace) -> int:
    from .edinet_client import EdinetClient

    cfg = load_pipeline_config(root)
    edinet_cfg = cfg.model_config.get("edinet", {})
    client = EdinetClient(
        base_url=edinet_cfg.get("base_url", "https://api.edinet-fsa.go.jp/api/v2"),
        timeout=int(edinet_cfg.get("request_timeout_seconds", 30)),
        retry_count=int(edinet_cfg.get("retry_count", 3)),
        retry_backoff_seconds=int(edinet_cfg.get("retry_backoff_seconds", 2)),
    )
    fiscal_years = args.fiscal_years or _configured_fiscal_years(cfg.company_year_master)
    dates = _annual_index_dates(cfg.company_master, fiscal_years)
    rows = _collect_index_rows(client, root, dates)
    written = write_table(root / "data" / "intermediate" / "document_index.parquet", rows)
    print(f"wrote {written} rows={len(rows)} fiscal_years={','.join(str(y) for y in fiscal_years)} dates={len(dates)}")
    return 0


def cmd_resolve(root: Path, args: argparse.Namespace) -> int:
    cfg = load_pipeline_config(root)
    docs = read_table(prefer_existing_table(root / "data" / "intermediate" / "document_index.parquet"))
    targets = resolve_target_documents(
        docs,
        cfg.company_master,
        cfg.company_year_master,
        cfg.document_filter,
        fiscal_years=args.fiscal_years,
        period_type=args.period_type,
    )
    if getattr(args, "merge", False):
        existing_path = prefer_existing_table(root / "data" / "intermediate" / "target_documents.parquet")
        existing = read_table(existing_path) if existing_path.exists() else []
        targets = _merge_rows_by_key(existing, targets, "company_year_id")
    written = write_table(root / "data" / "intermediate" / "target_documents.parquet", targets)
    print(f"wrote {written} rows={len(targets)}")
    return 0


def cmd_download(root: Path, args: argparse.Namespace) -> int:
    from .downloader import download_target_documents
    from .edinet_client import EdinetClient

    cfg = load_pipeline_config(root)
    edinet_cfg = cfg.model_config.get("edinet", {})
    client = EdinetClient(
        base_url=edinet_cfg.get("base_url", "https://api.edinet-fsa.go.jp/api/v2"),
        timeout=int(edinet_cfg.get("request_timeout_seconds", 30)),
        retry_count=int(edinet_cfg.get("retry_count", 3)),
        retry_backoff_seconds=int(edinet_cfg.get("retry_backoff_seconds", 2)),
    )
    targets = read_table(prefer_existing_table(root / args.target_documents))
    manifest = download_target_documents(client, targets, root / "data" / "raw" / "documents")
    written = write_table(root / "data" / "raw" / "download_manifest.parquet", manifest)
    print(f"wrote {written} rows={len(manifest)}")
    return 0


def cmd_build_edinet_db(root: Path, args: argparse.Namespace) -> int:
    db_path = _project_path(root, args.db)
    counts = build_edinet_db(root, db_path)
    print(f"wrote {db_path}")
    for key in sorted(counts):
        print(f"{key}: {counts[key]}")
    return 0


def cmd_extract_from_db(root: Path, args: argparse.Namespace) -> int:
    db_path = _project_path(root, args.db)
    output_path = _project_path(root, args.output)
    counts = extract_from_edinet_db(
        root,
        db_path,
        output_path,
        write_pipeline=not args.no_pipeline,
        period_type=getattr(args, "period_type", "annual"),
    )
    print(f"wrote {output_path}")
    for key in sorted(counts):
        print(f"{key}: {counts[key]}")
    return 0


def cmd_run_local(root: Path, args: argparse.Namespace) -> int:
    input_error = _local_run_input_error(root)
    if input_error:
        print(input_error)
        return 2
    db_path = _project_path(root, args.db)
    counts = build_edinet_db(root, db_path)
    print(f"wrote {db_path}")
    print(f"xbrl_facts: {counts.get('xbrl_facts', 0)}")
    extract_counts = extract_from_edinet_db(
        root,
        db_path,
        root / "data" / "intermediate" / "db_extracted_long.csv",
        write_pipeline=True,
        period_type="annual",
    )
    print(f"db_extracted_rows: {extract_counts.get('combined_rows', 0)}")
    cmd_import_manual_technicians(root, argparse.Namespace(note_path=""))
    cmd_normalize(root, argparse.Namespace())
    cmd_validate(root, argparse.Namespace())
    cmd_corroborate(root, argparse.Namespace())
    cmd_build_review_queue(root, argparse.Namespace(existing_excel=""))
    split_counts = split_local_review_rows(root)
    print(f"local_auto_accepted: {split_counts.get('accepted_rows', 0)}")
    print(f"local_manual_rows: {split_counts.get('manual_rows', 0)}")
    cmd_export_final(root, argparse.Namespace(reviewed="data/review/review_resolved.csv"))
    cmd_build_analysis(root, argparse.Namespace())
    cmd_report(root, argparse.Namespace())
    return 0


def cmd_run_all(root: Path, args: argparse.Namespace) -> int:
    if _local_run_input_error(root):
        fiscal_years = args.fiscal_years or list(range(2015, 2025))
        print("入力データが不足しているため、EDINET取得から実行します。")
        year_args = argparse.Namespace(fiscal_years=fiscal_years)
        for command in [cmd_index_annual, cmd_resolve]:
            code = command(root, year_args)
            if code:
                return code
        code = cmd_download(root, argparse.Namespace(target_documents="data/intermediate/target_documents.csv"))
        if code:
            return code
        code = cmd_locate_sections(root, argparse.Namespace())
        if code:
            return code
    return cmd_run_local(root, argparse.Namespace(db=args.db))


def cmd_extract_xbrl(root: Path, args: argparse.Namespace) -> int:
    cfg = load_pipeline_config(root)
    run_id = _run_id()
    targets = read_table(prefer_existing_table(root / "data" / "intermediate" / "target_documents.parquet"))
    rows: List[Dict[str, Any]] = []
    for target in targets:
        if target.get("resolution_status") != "resolved":
            continue
        if not _target_period_matches(target, getattr(args, "period_type", "annual")):
            continue
        csv_zip = root / "data" / "raw" / "documents" / str(target.get("docID")) / "csv.zip"
        rows.extend(extract_xbrl_csv_long(csv_zip, cfg.field_definition, target, run_id))
    written = write_table(root / "data" / "intermediate" / "xbrl_extracted_long.parquet", rows)
    print(f"wrote {written} rows={len(rows)}")
    return 0


def cmd_build_xbrl_fact_store(root: Path, args: argparse.Namespace) -> int:
    result = build_xbrl_fact_store(root, doc_id=args.doc_id, company_year_id=args.company_year_id, merge_existing=getattr(args, "merge", False))
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def cmd_extract_from_xbrl_fact_store(root: Path, args: argparse.Namespace) -> int:
    output_path = _project_path(root, args.output)
    result = extract_from_xbrl_fact_store(root, output_path, write_pipeline=args.write_pipeline, company_year_id=getattr(args, "company_year_id", ""))
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def cmd_compare_xbrl_fact_store(root: Path, args: argparse.Namespace) -> int:
    result = compare_xbrl_fact_store(
        root,
        _project_path(root, args.old),
        _project_path(root, args.new),
        _project_path(root, args.output),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def cmd_locate_sections(root: Path, args: argparse.Namespace) -> int:
    cfg = load_pipeline_config(root)
    run_id = _run_id()
    targets = read_table(prefer_existing_table(root / "data" / "intermediate" / "target_documents.parquet"))
    rows: List[Dict[str, Any]] = []
    for target in targets:
        if target.get("resolution_status") != "resolved":
            continue
        doc_dir = root / "data" / "raw" / "documents" / str(target.get("docID"))
        _extract_zip_text_payloads(doc_dir / "xbrl.zip", doc_dir / "extracted_xbrl")
        rows.extend(locate_candidate_blocks(doc_dir, target, cfg.extraction_sections, run_id))
    written = write_table(root / "data" / "intermediate" / "candidate_blocks.jsonl", rows)
    print(f"wrote {written} rows={len(rows)}")
    return 0


def cmd_normalize(root: Path, args: argparse.Namespace) -> int:
    cfg = load_pipeline_config(root)
    fields = field_map(cfg.field_definition)
    rows: List[Dict[str, Any]] = []
    for path in [
        prefer_existing_table(root / "data" / "intermediate" / "xbrl_extracted_long.parquet"),
        root / "data" / "intermediate" / "semiannual_h1_extracted_long.csv",
        root / "data" / "intermediate" / "manual_technician_extracted_long.csv",
        root / "data" / "intermediate" / "ai_extracted_long.jsonl",
    ]:
        if path.exists():
            rows.extend(read_table(path))
    normalized: List[Dict[str, Any]] = []
    for row in rows:
        field = fields.get(str(row.get("field_id")))
        if not field:
            copied = dict(row)
            copied["review_required"] = True
            copied["review_reason"] = "field_definition_missing"
            normalized.append(copied)
            continue
        normalized.append(normalize_extraction(row, field))
    written = write_table(root / "data" / "intermediate" / "normalized_long.parquet", normalized)
    print(f"wrote {written} rows={len(normalized)}")
    return 0


def cmd_validate(root: Path, args: argparse.Namespace) -> int:
    cfg = load_pipeline_config(root)
    rows = read_table(prefer_existing_table(root / "data" / "intermediate" / "normalized_long.parquet"))
    results = validate_records(rows, cfg.validation_rules)
    rows_with_status = attach_validation_status(rows, results)
    validation_path = write_table(root / "data" / "intermediate" / "validation_results.parquet", results)
    normalized_path = write_table(root / "data" / "intermediate" / "normalized_validated_long.parquet", rows_with_status)
    print(f"wrote {validation_path} rows={len(results)}")
    print(f"wrote {normalized_path} rows={len(rows_with_status)}")
    return 0


def cmd_corroborate(root: Path, args: argparse.Namespace) -> int:
    from .services import semantics_corroborate

    result = semantics_corroborate.run_corroboration(root)
    summary = result.get("summary", {})
    print(f"wrote {result.get('semantics_db_path')}")
    print(f"wrote {result.get('normalized_long_path')}")
    for key in sorted(summary):
        print(f"{key}={summary[key]}")
    print(f"auto_confirmed={summary.get('auto_confirmed', 0)}")
    print(f"conflicted={summary.get('conflicted', 0)}")
    return 0


def cmd_build_review_queue(root: Path, args: argparse.Namespace) -> int:
    cfg = load_pipeline_config(root)
    rows = _read_validated_or_normalized(root)
    existing = read_table(Path(args.existing_excel)) if args.existing_excel else []
    exclusions_path = root / "config" / "company_field_exclusions.csv"
    exclusions = read_table(exclusions_path) if exclusions_path.exists() else []
    cell_resolutions = _load_cell_resolutions_for_review_queue(root, cfg)
    queue = build_review_queue(
        rows, cfg.field_definition, _annual_company_years(cfg.company_year_master), existing, exclusions, cell_resolutions
    )
    written = write_table(root / "data" / "review" / "review_queue.xlsx", queue)
    write_table(root / "data" / "review" / "review_queue.csv", queue)
    print(f"wrote {written} rows={len(queue)}")
    return 0


def _load_cell_resolutions_for_review_queue(root: Path, cfg) -> Dict[tuple, Dict[str, Any]]:
    """corroboration.enable_review_downgrade が有効な場合のみ semantics.db から読む。

    安全スイッチ: config/validation_rules.yml の corroboration.enable_review_downgrade
    が False（明示的に無効化）の場合は None を返し、review_queue.build_review_queue の
    証拠ベース降格ロジックを完全にバイパスする。
    """
    policy = cfg.validation_rules.get("corroboration") or {}
    if policy.get("enable_review_downgrade") is False:
        return {}
    from .services import semantics_store

    db_path = semantics_store.semantics_db_path(root)
    if not db_path.exists():
        return {}
    conn = semantics_store.connect(root)
    try:
        raw = semantics_store.fetch_cell_resolutions(conn)
    finally:
        conn.close()
    # raw キーは (company_year_id, concept_id) = review_queue が期待する (company_year_id, field_id) と同一
    return raw


def cmd_split_local_review(root: Path, args: argparse.Namespace) -> int:
    counts = split_local_review_rows(root)
    print("wrote data/review/review_resolved_local_pass.xlsx")
    print("wrote data/review/review_queue_local_needs_manual.xlsx")
    for key in sorted(counts):
        print(f"{key}: {counts[key]}")
    return 0


def cmd_import_manual_technicians(root: Path, args: argparse.Namespace) -> int:
    note_path = Path(args.note_path).expanduser().resolve() if getattr(args, "note_path", "") else None
    summary = import_manual_technicians(root, note_path=note_path)
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def cmd_export_final(root: Path, args: argparse.Namespace) -> int:
    cfg = load_pipeline_config(root)
    annual_company_years = _annual_company_years(cfg.company_year_master)
    extracted = _read_validated_or_normalized(root)
    reviewed_path = prefer_existing_table(root / args.reviewed)
    reviewed = read_table(reviewed_path) if reviewed_path.exists() else []
    final_long = apply_review_decisions(extracted, reviewed)
    exportable = filter_exportable_rows(final_long)
    # P2: normalized_validated_long由来のcorroboration_status列をexporter側の
    # 列名resolutionへ明示マッピングする（apply_review_decisions自体は変更しない）。
    for row in exportable:
        row["resolution"] = row.get("corroboration_status", "")
    long_path = write_table(root / "data" / "final" / "final_master_long.parquet", exportable)
    long_csv_path = write_table(root / "data" / "final" / "final_master_long.csv", exportable)
    wide = build_wide_values(exportable, cfg.company_year_master, cfg.field_definition)
    wide_path = write_table(root / "data" / "final" / "wide_values.xlsx", wide)
    audit = build_source_audit(exportable, cfg.field_definition)
    audit_path = write_table(root / "data" / "final" / "source_audit.xlsx", audit)
    write_table(root / "data" / "final" / "wide_values.csv", wide)
    write_table(root / "data" / "final" / "final_master_wide.csv", wide)
    write_table(root / "data" / "final" / "source_audit.csv", audit)
    final_wide_path = root / "data" / "final" / "final_master_wide.xlsx"
    _write_final_workbook(final_wide_path, wide, audit, cfg.field_definition, annual_company_years)
    print(f"wrote {long_path} rows={len(exportable)}")
    print(f"wrote {long_csv_path} rows={len(exportable)}")
    print(f"wrote {wide_path} rows={len(wide)}")
    print(f"wrote {audit_path}")
    print(f"wrote {final_wide_path}")
    return 0


def _annual_company_years(company_years: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [row for row in company_years if str(row.get("period_type") or "annual") == "annual"]


def _company_year_period_type(company_year_id: str, annual_company_years: Iterable[Dict[str, Any]]) -> str:
    annual_ids = {str(row.get("company_year_id") or "") for row in annual_company_years}
    if company_year_id in annual_ids:
        return "annual"
    return "non_annual"


def _merge_rows_by_key(existing_rows: Iterable[Dict[str, Any]], new_rows: Iterable[Dict[str, Any]], key_name: str) -> List[Dict[str, Any]]:
    merged: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    for row in existing_rows:
        key = str(row.get(key_name) or "")
        if not key:
            continue
        if key not in merged:
            order.append(key)
        merged[key] = dict(row)
    for row in new_rows:
        key = str(row.get(key_name) or "")
        if not key:
            continue
        if key not in merged:
            order.append(key)
        merged[key] = dict(row)
    return [merged[key] for key in order]


def _target_period_matches(target: Dict[str, Any], period_type: str) -> bool:
    requested = str(period_type or "annual")
    if requested == "all":
        return True
    return str(target.get("period_type") or "annual") == requested


def cmd_build_analysis(root: Path, args: argparse.Namespace) -> int:
    wide_path = prefer_existing_table(root / "data" / "final" / "wide_values.xlsx")
    wide = read_table(wide_path)
    analysis = build_analysis_dataset(wide)
    written = write_table(root / "data" / "final" / "analysis_dataset.xlsx", analysis)
    csv_written = write_table(root / "data" / "final" / "analysis_dataset.csv", analysis)
    print(f"wrote {written} rows={len(analysis)}")
    print(f"wrote {csv_written} rows={len(analysis)}")
    return 0


def cmd_report(root: Path, args: argparse.Namespace) -> int:
    cfg = load_pipeline_config(root)
    run_id = _run_id()
    targets = _read_optional(root / "data" / "intermediate" / "target_documents.parquet")
    rows = _read_optional(root / "data" / "intermediate" / "normalized_validated_long.parquet")
    validations = _read_optional(root / "data" / "intermediate" / "validation_results.parquet")
    queue = _read_optional(root / "data" / "review" / "review_queue.xlsx")
    report = build_run_report(run_id, targets, rows, validations, queue)
    path = write_run_report(root / "data" / "final" / "run_report.md", report)
    coverage = _build_field_coverage(root, cfg.field_definition, cfg.company_year_master)
    coverage_path = write_table(root / "data" / "final" / "field_coverage.csv", coverage)
    markdown_path = root / "data" / "final" / "field_coverage.md"
    markdown_path.write_text(_field_coverage_markdown(coverage), encoding="utf-8")
    bundle_files = build_ai_bundle(root)
    print(f"wrote {path}")
    print(f"wrote {coverage_path}")
    print(f"wrote {markdown_path}")
    print(f"wrote {root / 'data' / 'ai_bundle'} files={len(bundle_files) + 2}")
    return 0


def cmd_build_ai_bundle(root: Path, args: argparse.Namespace) -> int:
    copied = build_ai_bundle(root)
    print(f"wrote {root / 'data' / 'ai_bundle'} files={len(copied) + 2}")
    return 0


def cmd_build_algorithm_audit(root: Path, args: argparse.Namespace) -> int:
    result = build_algorithm_audit_bundle(root)
    summary = result.get("summary", {})
    print(f"wrote {result.get('bundle_dir')} files={len(result.get('files', []))}")
    print(f"risk_flags: {summary.get('risk_flags', 0)}")
    print(f"review_derived_sections: {summary.get('review_derived_sections', 0)}")
    return 0


def cmd_build_corroboration_report(root: Path, args: argparse.Namespace) -> int:
    from .services.corroboration_report import build_corroboration_report

    result = build_corroboration_report(root)
    summary = result.get("summary", {})
    print(f"wrote {result.get('cells_path')}")
    print(f"cells_total={summary.get('cells_total', 0)}")
    print(f"corroborated_2plus={summary.get('corroborated_2plus', 0)}")
    print(f"corroborated_1={summary.get('corroborated_1', 0)}")
    print(f"corroborated_0={summary.get('corroborated_0', 0)}")
    print(f"conflicts={summary.get('conflicts', 0)}")
    return 0


def cmd_audit_findings(root: Path, args: argparse.Namespace) -> int:
    from .services.algorithm_audit_findings import build_algorithm_audit_findings

    result = build_algorithm_audit_findings(root)
    summary = result.get("summary", {})
    print(f"wrote {result.get('json_path')}")
    print(f"total_findings={summary.get('total', 0)}")
    for kind, count in (summary.get("by_kind") or {}).items():
        print(f"  {kind}={count}")
    return 0


def cmd_golden_freeze(root: Path, args: argparse.Namespace) -> int:
    from .services import golden

    result = golden.freeze_golden(root)
    print(f"golden_cell_count={result.get('golden_cell_count', 0)}")
    print(f"negative_golden_count={result.get('negative_golden_count', 0)}")
    for origin, count in sorted((result.get("by_origin") or {}).items()):
        print(f"by_origin.{origin}={count}")
    return 0


def cmd_backfill_semantics(root: Path, args: argparse.Namespace) -> int:
    from .services import semantics_backfill

    result = semantics_backfill.backfill_semantics(root)
    print(f"observed_items_total={result.get('observed_items_total', 0)}")
    for kind, count in sorted((result.get("observed_items_by_kind") or {}).items()):
        print(f"observed_items_by_kind.{kind}={count}")
    print(f"concept_mappings_total={result.get('concept_mappings_total', 0)}")
    for action, count in sorted((result.get("concept_mappings_by_action") or {}).items()):
        print(f"concept_mappings_by_action.{action}={count}")
    return 0


def cmd_semantics_coverage(root: Path, args: argparse.Namespace) -> int:
    from .services import semantics_coverage

    result = semantics_coverage.build_and_write_coverage_report(root)
    print(f"wrote {result.get('report_json_path')}")
    summary = result.get("summary", {})
    print(f"observed_total={summary.get('observed_total', 0)}")
    print(f"observed_unmapped={summary.get('observed_unmapped', 0)}")
    print(f"mappings_total={summary.get('mappings_total', 0)}")
    return 0


def cmd_infer_sources_from_confirmed(root: Path, args: argparse.Namespace) -> int:
    """BuildBase S1a: 出典逆引きエンジンのdry-runレポートを生成する（読み取り専用）。

    final/review/config/semantics.db/edinet.db への書き込みは一切行わない。
    出力は data/reports/source_inference_dry_run.json / .md のみ。
    """
    from .services import source_inference

    field_ids = [field.strip() for field in str(args.fields).split(",") if field.strip()]
    result = source_inference.build_dry_run_report(root, field_ids)
    report = result["report"]
    repro = report.get("reproduction", {})
    print(f"wrote {result.get('report_json_path')}")
    print(f"known_cells_total={repro.get('total', 0)}")
    print(f"known_cells_matched_high_confidence={repro.get('matched_high_confidence', 0)}")
    print(f"known_cells_reproduction_rate={repro.get('rate', 0.0):.3f}")
    for status, count in sorted((report.get("summary") or {}).items()):
        print(f"missing_cells.{status}={count}")
    return 0


def cmd_apply_source_inference(root: Path, args: argparse.Namespace) -> int:
    """BuildBase S1b: 出典逆引きの学習パターンを会社×fieldの複数年度成立ゲート付きで適用する。

    既定は --dry-run（何も書き込まない）。--apply 明示時のみ
    reviews.upsert_resolved_reviews 経由で review_resolved.csv に書き込み、
    export-final 以降を再実行して final に反映する。
    既存の値があるセル・既にレビュー確定済みのセルには一切書き込まない。
    """
    from .services import source_inference

    field_ids = [field.strip() for field in str(args.fields).split(",") if field.strip()]
    plan = source_inference.build_promotion_plan(root, field_ids)
    result = source_inference.apply_promotion_plan(root, plan, dry_run=bool(args.dry_run))

    print(f"dry_run={result.get('dry_run')}")
    print(f"promote_candidates={len(plan.get('promote') or [])}")
    print(f"candidate_single_year={len(plan.get('candidate_single_year') or [])}")
    print(f"suspect_existing_values={len(plan.get('suspect_existing_values') or [])}")
    print(f"planned={result.get('planned', 0)}")
    print(f"skipped_existing_review={result.get('skipped_existing_review', 0)}")
    print(f"applied={result.get('applied', 0)}")
    if not result.get("dry_run"):
        apply_review_result = result.get("apply_review") or {}
        print(f"apply_review_ran={apply_review_result.get('ran')}")
        if apply_review_result.get("ran"):
            print(f"apply_review_exit_code={apply_review_result.get('exit_code')}")
    return 0


def cmd_ai_map_observed_items(root: Path, args: argparse.Namespace) -> int:
    from .ai_runner import ClaudeCliRunner
    from .services import ai_mapping

    dry_run = bool(args.dry_run)
    runner = None if dry_run else ClaudeCliRunner()
    result = ai_mapping.run_ai_mapping_batch(
        root, runner=runner, tier=args.tier, limit=args.limit, dry_run=dry_run,
        taxonomy_kind=getattr(args, "taxonomy_kind", None),
    )
    print(f"observed_items_targeted={result.get('observed_items_targeted', 0)}")
    print(f"chunks={result.get('chunks', 0)}")
    print(f"ai_calls_made={result.get('ai_calls_made', 0)}")
    print(f"proposals_written={result.get('proposals_written', 0)}")
    print(f"parse_errors={result.get('parse_errors', 0)}")
    print(f"dry_run={result.get('dry_run')}")
    return 0


def cmd_ai_verify_mappings(root: Path, args: argparse.Namespace) -> int:
    from .services import ai_mapping

    result = ai_mapping.verify_ai_proposals_against_corroboration(root)
    print(f"ai_proposed_total={result.get('ai_proposed_total', 0)}")
    print(f"likely_confirmable_via_corroboration={result.get('likely_confirmable_via_corroboration', 0)}")
    return 0


def cmd_promote_ai_mappings(root: Path, args: argparse.Namespace) -> int:
    """P5c: AI提案(map/new_concept)の数値裏取り確定・新概念採用を実行する。

    --confirm-verified-maps / --adopt-new-concepts のいずれも未指定の場合は
    両方を実行する（既定挙動）。どちらか一方のみ指定した場合はそちらのみ実行。
    既定は --dry-run（DB書き込みなし、判定結果の表示のみ）。
    """
    from .services import mapping_promotion

    dry_run = bool(args.dry_run)
    run_maps = getattr(args, "confirm_verified_maps", None)
    run_new_concepts = getattr(args, "adopt_new_concepts", None)
    if run_maps is None and run_new_concepts is None:
        run_maps = True
        run_new_concepts = True
    else:
        run_maps = bool(run_maps)
        run_new_concepts = bool(run_new_concepts)

    if run_maps:
        map_result = mapping_promotion.promote_verified_map_proposals(root, dry_run=dry_run)
        print(f"map_proposals_checked={map_result.get('map_proposals_checked', 0)}")
        print(f"map_corroborated={map_result.get('corroborated', 0)}")
        print(f"map_not_corroborated={map_result.get('not_corroborated', 0)}")

    if run_new_concepts:
        concept_result = mapping_promotion.adopt_new_concepts(root, dry_run=dry_run)
        print(f"new_concept_proposals_checked={concept_result.get('new_concept_proposals_checked', 0)}")
        print(f"concepts_created={concept_result.get('concepts_created', 0)}")
        print(f"new_concept_mappings_confirmed={concept_result.get('mappings_confirmed', 0)}")

    print(f"dry_run={dry_run}")
    return 0


def cmd_regression_check(root: Path, args: argparse.Namespace) -> int:
    from .services import golden

    summary = golden.run_regression(root, mode=args.mode)
    for key in sorted(summary):
        print(f"{key}={summary[key]}")
    return 0 if summary.get("pass") else 1


def cmd_enrich_field_definition(root: Path, args: argparse.Namespace) -> int:
    """P4b: field_definition.csv/.xlsx の増分エンリッチ（確定AI mapの追記＋新概念行追加）。

    既定は --dry-run（計画表示のみ、書き込みなし）。--apply で実際に書き込む。
    """
    from .services import field_definition_enrich

    dry_run = bool(args.dry_run)
    plan = field_definition_enrich.build_enrichment_plan(root)
    print(f"dry_run={dry_run}")
    print(f"append_candidates={len(plan['appends'])}")
    print(f"append_count(with_changes)={plan['append_count']}")
    for item in plan["appends"]:
        if item["tags_to_add"]:
            print(f"  append field_id={item['field_id']} tags_to_add={';'.join(item['tags_to_add'])}")
    print(f"new_row_count={plan['new_row_count']}")
    for row in plan["new_rows"]:
        print(f"  new_row field_id={row['field_id']} field_name_ja={row['field_name_ja']} xbrl_tag_candidates={row['xbrl_tag_candidates']}")

    if dry_run:
        return 0

    result = field_definition_enrich.apply_enrichment(root)
    print(f"appended_count={result['appended_count']}")
    print(f"added_count={result['added_count']}")
    return 0


def cmd_automation_status(root: Path, args: argparse.Namespace) -> int:
    from .services import automation

    result = automation.automation_status(root, fiscal_year=args.fiscal_year)
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def cmd_roll_forward_year(root: Path, args: argparse.Namespace) -> int:
    from .services import automation

    result = automation.roll_forward_company_years(root, args.fiscal_year, dry_run=args.dry_run)
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def cmd_annual_refresh(root: Path, args: argparse.Namespace) -> int:
    from .services import pipeline

    return pipeline.annual_refresh(
        root,
        fiscal_year=args.fiscal_year,
        force=args.force,
        dry_run=args.dry_run,
        log=print,
    )


def cmd_stock_status(root: Path, args: argparse.Namespace) -> int:
    from .services import market

    result = market.stock_monthly_status(root)
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def cmd_stock_refresh(root: Path, args: argparse.Namespace) -> int:
    from .services import market

    start_date = _parse_date(args.start_date) if args.start_date else None
    end_date = _parse_date(args.end_date) if args.end_date else None
    if start_date or end_date:
        result = market.refresh_stock_prices(
            root,
            start_date=start_date,
            end_date=end_date,
            force=True,
            dry_run=args.dry_run,
            log=print,
        )
    else:
        result = market.refresh_stock_prices_if_due(root, force=args.force, dry_run=args.dry_run, log=print)
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0 if str(result.get("status")) in {"succeeded", "partial_success", "dry_run", "skipped"} else 1


def cmd_factbook_status(root: Path, args: argparse.Namespace) -> int:
    from .services import company_factbooks

    result = company_factbooks.factbook_status(root)
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def cmd_factbook_refresh(root: Path, args: argparse.Namespace) -> int:
    from .ai_runner import ClaudeCliRunner
    from .services import company_factbooks

    runner = ClaudeCliRunner() if bool(getattr(args, "use_ai", False)) and not bool(args.dry_run) else None
    result = company_factbooks.refresh_company_factbooks(
        root,
        force=args.force,
        dry_run=args.dry_run,
        log=print,
        ai_runner=runner,
        ai_tier=str(getattr(args, "ai_tier", "bulk") or "bulk"),
        company_ids=getattr(args, "company", []),
        source_ids=getattr(args, "source", []),
        fiscal_years=getattr(args, "fiscal_year", []),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0 if str(result.get("status")) in {"succeeded", "partial_success", "dry_run"} else 1


def cmd_factbook_validate(root: Path, args: argparse.Namespace) -> int:
    from .services import company_factbooks

    result = company_factbooks.validate_factbook_against_yuho(root)
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0 if str(result.get("status")) in {"completed", "incomplete"} else 1


def cmd_factbook_coverage(root: Path, args: argparse.Namespace) -> int:
    from .services import company_factbooks

    result = company_factbooks.build_factbook_target_coverage(root)
    print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def cmd_init_xlsx(root: Path, args: argparse.Namespace) -> int:
    config_dir = root / "config"
    for name in ["company_master", "company_year_master", "field_definition"]:
        rows = read_table(config_dir / f"{name}.csv")
        written = write_table(config_dir / f"{name}.xlsx", rows)
        print(f"wrote {written}")
    return 0


def cmd_web(root: Path, args: argparse.Namespace) -> int:
    os.environ.setdefault("YUHO_PROJECT_ROOT", str(root))
    from .web_api.__main__ import main as web_main

    argv = ["--host", args.host, "--port", str(args.port)]
    if args.reload:
        argv.append("--reload")
    return web_main(argv)


def _read_validated_or_normalized(root: Path) -> List[Dict[str, Any]]:
    validated = prefer_existing_table(root / "data" / "intermediate" / "normalized_validated_long.parquet")
    if validated.exists():
        return read_table(validated)
    return read_table(prefer_existing_table(root / "data" / "intermediate" / "normalized_long.parquet"))


def _build_field_coverage(
    root: Path,
    field_definitions: List[Dict[str, Any]],
    company_year_master: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    wide_path = root / "data" / "final" / "wide_values.csv"
    wide_rows = read_table(wide_path) if wide_path.exists() else []
    total = len(wide_rows) if wide_rows else len(company_year_master)
    if wide_rows:
        counts = {
            str(field.get("field_id") or ""): sum(1 for row in wide_rows if _has_output_value(row.get(str(field.get("field_id") or ""))))
            for field in field_definitions
            if field.get("field_id")
        }
    else:
        final_long = _read_optional(root / "data" / "final" / "final_master_long.parquet")
        counts = {}
        for row in final_long:
            if _has_output_value(row.get("value")):
                counts[str(row.get("field_id") or "")] = counts.get(str(row.get("field_id") or ""), 0) + 1
    coverage = []
    for field in field_definitions:
        field_id = str(field.get("field_id") or "")
        filled = counts.get(field_id, 0)
        coverage.append(
            {
                "field_id": field_id,
                "field_name_ja": field.get("field_name_ja", ""),
                "category": field.get("category", ""),
                "preferred_method": field.get("preferred_method", ""),
                "filled_company_years": filled,
                "total_company_years": total,
                "coverage_pct": round(filled / total, 4) if total else 0,
            }
        )
    return sorted(coverage, key=lambda row: (-float(row["filled_company_years"]), str(row["field_id"])))


def _field_coverage_markdown(rows: List[Dict[str, Any]]) -> str:
    total_fields = len(rows)
    total_company_years = rows[0].get("total_company_years", 0) if rows else 0
    filled_fields = sum(1 for row in rows if int(row.get("filled_company_years") or 0) > 0)
    lines = [
        "# Field Coverage",
        "",
        f"- fields: {total_fields}",
        f"- company_years: {total_company_years}",
        f"- filled_fields: {filled_fields}",
        "",
        "| field_id | field_name_ja | filled | coverage |",
        "| --- | --- | ---: | ---: |",
    ]
    for row in rows:
        pct = float(row.get("coverage_pct") or 0) * 100
        lines.append(
            f"| {row.get('field_id', '')} | {row.get('field_name_ja', '')} | {row.get('filled_company_years', 0)} | {pct:.1f}% |"
        )
    return "\n".join(lines) + "\n"


def _has_output_value(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return text not in {"", "nan", "NaN", "None", "null"}


def _local_run_input_error(root: Path) -> str:
    target_path = prefer_existing_table(root / "data" / "intermediate" / "target_documents.parquet")
    candidate_path = root / "data" / "intermediate" / "candidate_blocks.jsonl"
    document_dir = root / "data" / "raw" / "documents"
    if not target_path.exists() or not read_table(target_path):
        return (
            "入力データが不足しています: data/intermediate/target_documents がありません。\n"
            "先に index-annual、resolve、download、locate-sections を実行してください。"
        )
    if not document_dir.exists() or not any(document_dir.glob("*/csv.zip")):
        return (
            "入力データが不足しています: data/raw/documents/*/csv.zip がありません。\n"
            "先に download を実行してください。"
        )
    if not candidate_path.exists() or candidate_path.stat().st_size == 0:
        return (
            "入力データが不足しています: data/intermediate/candidate_blocks.jsonl がありません。\n"
            "先に locate-sections を実行してください。"
        )
    return ""


def _read_optional(path: Path) -> List[Dict[str, Any]]:
    actual = prefer_existing_table(path)
    return read_table(actual) if actual.exists() else []


def _project_path(root: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else root / path


def _load_env_file(root: Path) -> None:
    for name in [".env", ".env.txt", "env.txt"]:
        path = root / name
        if not path.exists():
            continue
        with path.open("r", encoding="utf-8-sig") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
        return


def _date_range(start: date, end: date) -> Iterable[date]:
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def _collect_index_rows(client: Any, root: Path, dates: Iterable[date]) -> List[Dict[str, Any]]:
    from .edinet_client import save_index_response

    rows: List[Dict[str, Any]] = []
    for current in dates:
        response = client.list_documents(current, doc_type=2)
        save_index_response(root / "data" / "raw" / "edinet_index" / f"{current.isoformat()}.json", response)
        for item in response.get("results", []):
            copied = dict(item)
            copied["fileDate"] = current.isoformat()
            rows.append(copied)
    return rows


def _configured_fiscal_years(company_years: List[Dict[str, Any]]) -> List[int]:
    return sorted({int(row["fiscal_year"]) for row in company_years})


def _annual_index_dates(company_master: List[Dict[str, Any]], fiscal_years: List[int]) -> List[date]:
    fiscal_months = {int(row.get("fiscal_year_end_month") or 3) for row in company_master}
    dates = set()
    for fiscal_year in fiscal_years:
        for month in fiscal_months:
            if month == 12:
                start = date(fiscal_year + 1, 3, 1)
                end = date(fiscal_year + 1, 4, 30)
            else:
                start = date(fiscal_year + 1, 6, 1)
                end = date(fiscal_year + 1, 7, 31)
            dates.update(_date_range(start, end))
    return sorted(dates)


def _parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def _run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ") + "-" + uuid.uuid4().hex[:8]


def _extract_zip_text_payloads(zip_path: Path, output_dir: Path) -> None:
    if not zip_path.exists():
        return
    output_dir.mkdir(parents=True, exist_ok=True)
    allowed = {".html", ".htm", ".xhtml", ".txt", ".csv"}
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            name = info.filename
            suffix = Path(name).suffix.lower()
            if suffix not in allowed:
                continue
            target = output_dir / Path(name).name
            if not target.exists():
                target.write_bytes(zf.read(info))


def _write_final_workbook(path: Path, wide: List[Dict[str, Any]], audit: List[Dict[str, Any]], fields: List[Dict[str, Any]], company_years: List[Dict[str, Any]]) -> None:
    write_table(path.with_name("final_master_wide_values.csv"), wide)
    write_table(path.with_name("final_master_source_audit.csv"), audit)
    write_table(path.with_name("final_master_field_definition.csv"), fields)
    write_table(path.with_name("final_master_company_year_master.csv"), company_years)
    try:
        import openpyxl  # type: ignore
    except ImportError:
        write_table(path, wide)
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    _write_sheet(wb.active, "wide_values", wide)
    _write_sheet(wb.create_sheet(), "source_audit", audit)
    _write_sheet(wb.create_sheet(), "field_definition", fields)
    _write_sheet(wb.create_sheet(), "company_year_master", company_years)
    _write_sheet(wb.create_sheet(), "run_summary", [{"sheet": "wide_values", "rows": len(wide)}, {"sheet": "source_audit", "rows": len(audit)}])
    wb.save(path)


def _write_sheet(ws: Any, title: str, rows: List[Dict[str, Any]]) -> None:
    ws.title = title[:31]
    if not rows:
        return
    headers: List[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(key, "") for key in headers])


if __name__ == "__main__":
    raise SystemExit(main())
